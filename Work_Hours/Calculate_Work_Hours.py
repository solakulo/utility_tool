import os
import json
import openpyxl
from datetime import datetime, timedelta
import sys

CONFIG_FILE = "work_schedule_config.json"
LOG_DIR = "logs"


def load_config():
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def read_logs(year):
    log_file = os.path.join(LOG_DIR, f"{year}.log")
    if not os.path.exists(log_file):
        return []  # 如果日志文件不存在，返回空列表

    logs = []
    with open(log_file, "r", encoding="utf-8") as f:
        for line in f:
            parts = line.strip().split(", ")
            if len(parts) == 3:
                date, first_time, last_time = parts
                logs.append(
                    {"date": date, "first_time": first_time, "last_time": last_time}
                )

    # 按日期升序排序
    logs.sort(key=lambda x: x["date"])
    return logs


def append_to_logs(new_logs, year):
    log_file = os.path.join(LOG_DIR, f"{year}.log")

    # 读取现有日志
    existing_logs = read_logs(year)
    existing_dates = {log["date"] for log in existing_logs}

    # 过滤新增日志
    filtered_logs = [log for log in new_logs if log["date"] not in existing_dates]

    # 如果有新增内容，追加到文件
    if filtered_logs:
        with open(log_file, "a", encoding="utf-8") as f:
            for log in filtered_logs:
                f.write(f"{log['date']}, {log['first_time']}, {log['last_time']}\n")

    return existing_logs + filtered_logs


def calculate_rest_duration(work_start, work_end, rest_periods):
    total_rest_duration = timedelta()
    for period in rest_periods:
        start, end = period.split("-")
        rest_start = datetime.strptime(start, "%H:%M")
        rest_end = datetime.strptime(end, "%H:%M")

        # 1.1 完整包括在上下班时间段之间
        if rest_start >= work_start and rest_end <= work_end:
            total_rest_duration += rest_end - rest_start

        # 1.2 完整在上下班时间之外 (不计算)
        elif rest_end <= work_start or rest_start >= work_end:
            continue

        # 1.3 与上下班时间段有交集
        else:
            overlap_start = max(work_start, rest_start)
            overlap_end = min(work_end, rest_end)
            if overlap_end > overlap_start:
                total_rest_duration += overlap_end - overlap_start

    return total_rest_duration


def calculate_work_hours(log, config):
    work_start = datetime.strptime(log["first_time"], "%H:%M:%S")
    work_end = datetime.strptime(log["last_time"], "%H:%M:%S")

    # 计算包含所有休息时间的上班时长
    total_work_duration = work_end - work_start

    # 计算当天有效的休息时间段
    rest_periods = config.get("rest_periods", [])
    total_rest_duration = calculate_rest_duration(work_start, work_end, rest_periods)

    # 去除休息的上班时长
    effective_work_duration = total_work_duration - total_rest_duration

    # 标准工作时长
    standard_duration = timedelta(
        hours=int(config["standard_work_hours"].split(":")[0]),
        minutes=int(config["standard_work_hours"].split(":")[1]),
    )

    # 计算加班时长
    overtime = max(timedelta(0), effective_work_duration - standard_duration)

    return total_work_duration, effective_work_duration, overtime


def save_to_excel(logs, config, year):
    output_file = f"system_logs_{year}.xlsx"

    # 加载或创建工作簿
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    # 按月份创建或选择工作表
    for month in range(1, 13):
        sheet_name = f"{month:02}"
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(
                [
                    "日期",
                    "最早开机时间",
                    "最晚关机时间",
                    "包含休息上班时长",
                    "去除休息上班时长",
                    "加班时长",
                ]
            )
        else:
            ws = wb[sheet_name]

        # 获取已有数据的日期
        existing_dates = {
            row[0].value for row in ws.iter_rows(min_row=2) if row[0].value
        }

        # 写入新增日志数据
        for log in sorted(logs, key=lambda x: x["date"]):  # 按日期升序排序
            if log["date"] not in existing_dates:
                total, effective, overtime = calculate_work_hours(log, config)
                ws.append(
                    [
                        log["date"],
                        log["first_time"],
                        log["last_time"],
                        str(total),
                        str(effective),
                        str(overtime),
                    ]
                )

    wb.save(output_file)
    print(f"生成的图表已保存到 {output_file}")


if __name__ == "__main__":
    year = sys.argv[1]
    config = load_config()
    new_logs = read_logs(year)
    updated_logs = append_to_logs(new_logs, year)
    save_to_excel(updated_logs, config, year)
